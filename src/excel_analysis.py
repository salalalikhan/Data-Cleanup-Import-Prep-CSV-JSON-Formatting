#!/usr/bin/env python3
"""
Creates a comprehensive Excel workbook with multiple sheets for data analysis
"""

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import json
from datetime import datetime
import os

def load_data():
    data = {}
    
    # Load main cleaned data
    if os.path.exists('outputs/cleaned_import.csv'):
        data['main'] = pd.read_csv('outputs/cleaned_import.csv')
        print(f"Loaded main data: {len(data['main'])} records")
    
    # Load data quality issues
    if os.path.exists('outputs/data_quality_issues.csv'):
        data['quality_issues'] = pd.read_csv('outputs/data_quality_issues.csv')
        print(f"Loaded quality issues: {len(data['quality_issues'])} issues")
    
    # Load field mapping
    if os.path.exists('outputs/field_mapping.csv'):
        data['field_mapping'] = pd.read_csv('outputs/field_mapping.csv')
        print(f"Loaded field mappings: {len(data['field_mapping'])} mappings")
    
    # Load JSON data if available
    if os.path.exists('outputs/cleaned_import.json'):
        with open('outputs/cleaned_import.json', 'r') as f:
            data['json_data'] = json.load(f)
        print(f"Loaded JSON data: {len(data['json_data'])} records")
    
    return data

def create_summary_sheet(wb, data):
    ws = wb.create_sheet("Summary")
    
    # Title
    ws['A1'] = "Data Analysis Summary"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    
    # Summary statistics
    if 'main' in data:
        df = data['main']
        
        # Basic stats
        ws['A3'] = "Dataset Overview"
        ws['A3'].font = Font(bold=True)
        
        ws['A4'] = "Total Records:"
        ws['B4'] = len(df)
        
        ws['A5'] = "Active Clients:"
        ws['B5'] = df['active_status'].sum() if 'active_status' in df.columns else 'N/A'
        
        ws['A6'] = "Total Balance (USD):"
        ws['B6'] = f"${df['balance_cents'].sum() / 100:,.2f}" if 'balance_cents' in df.columns else 'N/A'
        
        ws['A7'] = "Average Balance (USD):"
        ws['B7'] = f"${df['balance_cents'].mean() / 100:,.2f}" if 'balance_cents' in df.columns else 'N/A'
        
        # Data quality summary
        ws['A9'] = "Data Quality"
        ws['A9'].font = Font(bold=True)
        
        if 'quality_issues' in data:
            ws['A10'] = "Total Issues Found:"
            ws['B10'] = len(data['quality_issues'])
            
            # Issue breakdown by field
            issue_counts = data['quality_issues']['field'].value_counts()
            row = 11
            for field, count in issue_counts.items():
                ws[f'A{row}'] = f"  {field}:"
                ws[f'B{row}'] = count
                row += 1
    
    # Format the sheet
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.endswith(':'):
                cell.font = Font(bold=True)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_data_sheet(wb, data, sheet_name, df):
    """Create a formatted data sheet"""
    ws = wb.create_sheet(sheet_name)
    
    # Add data to worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Format header row
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Create table
    if len(df) > 0:
        tab = Table(displayName=sheet_name.replace(" ", "_"), ref=f"A1:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                              showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add conditional formatting for numeric columns
    if sheet_name == "Main Data" and 'balance_cents' in df.columns:
        balance_col = df.columns.get_loc('balance_cents') + 1
        col_letter = openpyxl.utils.get_column_letter(balance_col)
        ws.conditional_formatting.add(f'{col_letter}2:{col_letter}{ws.max_row}',
                                     ColorScaleRule(start_type='min', start_color='FF6B6B',
                                                   end_type='max', end_color='4ECDC4'))

def create_analytics_sheet(wb, data):
    """Create advanced analytics with formulas"""
    ws = wb.create_sheet("Analytics")
    
    if 'main' not in data:
        ws['A1'] = "No main data available for analytics"
        return
    
    df = data['main']
    
    # Title
    ws['A1'] = "Advanced Analytics"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:E1')
    
    # Client distribution by status
    ws['A3'] = "Client Status Distribution"
    ws['A3'].font = Font(bold=True)
    
    if 'active_status' in df.columns:
        active_count = df['active_status'].sum()
        inactive_count = len(df) - active_count
        
        ws['A4'] = "Active Clients:"
        ws['B4'] = active_count
        ws['C4'] = f"=B4/({active_count + inactive_count})"
        ws['C4'].number_format = '0.0%'
        
        ws['A5'] = "Inactive Clients:"
        ws['B5'] = inactive_count
        ws['C5'] = f"=B5/({active_count + inactive_count})"
        ws['C5'].number_format = '0.0%'
    
    # Balance analytics
    ws['A7'] = "Balance Analytics"
    ws['A7'].font = Font(bold=True)
    
    if 'balance_cents' in df.columns:
        ws['A8'] = "Total Balance:"
        ws['B8'] = df['balance_cents'].sum()
        ws['B8'].number_format = '$#,##0.00'
        
        ws['A9'] = "Average Balance:"
        ws['B9'] = f"=AVERAGE('Main Data'!G:G)"
        ws['B9'].number_format = '$#,##0.00'
        
        ws['A10'] = "Median Balance:"
        ws['B10'] = f"=MEDIAN('Main Data'!G:G)"
        ws['B10'].number_format = '$#,##0.00'
        
        ws['A11'] = "Max Balance:"
        ws['B11'] = f"=MAX('Main Data'!G:G)"
        ws['B11'].number_format = '$#,##0.00'
        
        ws['A12'] = "Min Balance:"
        ws['B12'] = f"=MIN('Main Data'!G:G)"
        ws['B12'].number_format = '$#,##0.00'

    ws['A14'] = "Company Analytics"
    ws['A14'].font = Font(bold=True)
    
    if 'company_id' in df.columns:
        company_counts = df['company_id'].value_counts()
        ws['A15'] = "Total Companies:"
        ws['B15'] = len(company_counts)
        
        ws['A16'] = "Avg Clients per Company:"
        ws['B16'] = f"=ROUND(COUNTA('Main Data'!I:I)/B15,2)"

def create_charts_sheet(wb, data):
    ws = wb.create_sheet("Charts")
    
    if 'main' not in data:
        ws['A1'] = "No data available for charts"
        return
    
    df = data['main']

    ws['A1'] = "Data Visualizations"
    ws['A1'].font = Font(size=16, bold=True)

    if 'active_status' in df.columns:
        # Active vs Inactive pie chart data
        ws['A3'] = "Status"
        ws['B3'] = "Count"
        ws['A4'] = "Active"
        ws['B4'] = df['active_status'].sum()
        ws['A5'] = "Inactive"
        ws['B5'] = len(df) - df['active_status'].sum()
        
        # Create pie chart
        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=4, max_row=5)
        data_ref = Reference(ws, min_col=2, min_row=3, max_row=5)
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Client Status Distribution"
        pie.height = 10
        pie.width = 15
        ws.add_chart(pie, "D3")

def create_pivot_summary_sheet(wb, data):
    ws = wb.create_sheet("Pivot Data")
    
    if 'main' not in data:
        ws['A1'] = "No data available for pivot analysis"
        return
    
    df = data['main']
    
    # Company summary for pivot tables
    if 'company_id' in df.columns and 'balance_cents' in df.columns:
        company_summary = df.groupby('company_id').agg({
            'record_id': 'count',
            'balance_cents': ['sum', 'mean'],
            'active_status': 'sum'
        }).round(2)
        
        company_summary.columns = ['Client_Count', 'Total_Balance', 'Avg_Balance', 'Active_Clients']
        company_summary = company_summary.reset_index()
        
        ws['A1'] = "Company Summary for Pivot Analysis"
        ws['A1'].font = Font(bold=True)
        
        for r in dataframe_to_rows(company_summary, index=False, header=True):
            ws.append(r)
    
        if len(company_summary) > 0:
            tab = Table(displayName="CompanySummary", ref=f"A2:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}")
            style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                  showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            tab.tableStyleInfo = style
            ws.add_table(tab)

def main():
    print("Creating comprehensive Excel analysis workbook...")
    
    # Load data
    data = load_data()
    
    if not data:
        print("No data files found. Please ensure cleaned data files exist in outputs/ directory.")
        return

    wb = Workbook()
    wb.remove(wb.active)
    create_summary_sheet(wb, data)
    
    if 'main' in data:
        create_data_sheet(wb, data, "Main Data", data['main'])
    
    if 'quality_issues' in data:
        create_data_sheet(wb, data, "Quality Issues", data['quality_issues'])
    
    if 'field_mapping' in data:
        create_data_sheet(wb, data, "Field Mappings", data['field_mapping'])
    
    create_analytics_sheet(wb, data)
    create_charts_sheet(wb, data)
    create_pivot_summary_sheet(wb, data)
    
    output_file = "outputs/data_analysis_workbook.xlsx"
    wb.save(output_file)
    print(f"Excel workbook created successfully: {output_file}")

    print(f"\nWorkbook contains {len(wb.sheetnames)} sheets:")
    for sheet in wb.sheetnames:
        print(f"  - {sheet}")

if __name__ == "__main__":
    main()