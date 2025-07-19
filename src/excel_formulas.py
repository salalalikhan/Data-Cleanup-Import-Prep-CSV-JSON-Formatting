#!/usr/bin/env python3
"""
Advanced Excel Formulas Enhancement
Adds sophisticated Excel formulas, functions, and data analysis features
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule, FormulaRule
from openpyxl.chart import BarChart, PieChart, LineChart, ScatterChart, Reference
from openpyxl.chart.trendline import Trendline
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
import pandas as pd

def add_advanced_analytics_sheet(wb):
    """Create advanced analytics with complex formulas"""
    ws = wb.create_sheet("Advanced Analytics", 7)
    
    # Title and header
    ws['A1'] = "Advanced Financial Analytics Dashboard"
    ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Key Performance Indicators section
    ws['A3'] = "Key Performance Indicators"
    ws['A3'].font = Font(size=14, bold=True, color="1F4E79")
    ws.merge_cells('A3:D3')
    
    # Revenue Analysis
    ws['A5'] = "Revenue Analysis"
    ws['A5'].font = Font(bold=True)
    
    ws['A6'] = "Total Portfolio Value:"
    ws['B6'] = "=SUM('Main Data'!G:G)/100"
    ws['B6'].number_format = '"$"#,##0.00'
    
    ws['A7'] = "Average Account Value:"
    ws['B7'] = "=AVERAGE('Main Data'!G:G)/100"
    ws['B7'].number_format = '"$"#,##0.00'
    
    ws['A8'] = "Median Account Value:"
    ws['B8'] = "=MEDIAN('Main Data'!G:G)/100"
    ws['B8'].number_format = '"$"#,##0.00'
    
    ws['A9'] = "Standard Deviation:"
    ws['B9'] = "=STDEV('Main Data'!G:G)/100"
    ws['B9'].number_format = '"$"#,##0.00'
    
    ws['A10'] = "Coefficient of Variation:"
    ws['B10'] = "=B9/B7"
    ws['B10'].number_format = "0.00%"
    
    # Client Segmentation Analysis
    ws['E5'] = "Client Segmentation"
    ws['E5'].font = Font(bold=True)
    
    ws['E6'] = "High Value (>$2000):"
    ws['F6'] = "=COUNTIF('Main Data'!G:G,\">200000\")"
    
    ws['E7'] = "Medium Value ($500-$2000):"
    ws['F7'] = "=COUNTIFS('Main Data'!G:G,\">=50000\",'Main Data'!G:G,\"<=200000\")"
    
    ws['E8'] = "Low Value (<$500):"
    ws['F8'] = "=COUNTIF('Main Data'!G:G,\"<50000\")"
    
    ws['E9'] = "% High Value Clients:"
    ws['F9'] = "=F6/COUNTA('Main Data'!G:G)"
    ws['F9'].number_format = "0.00%"
    
    # Risk Analysis
    ws['A12'] = "Risk Analysis"
    ws['A12'].font = Font(bold=True)
    
    ws['A13'] = "Clients with Zero Balance:"
    ws['B13'] = "=COUNTIF('Main Data'!G:G,0)"
    
    ws['A14'] = "Inactive High-Value Clients:"
    ws['B14'] = "=SUMPRODUCT(('Main Data'!E:E=0)*('Main Data'!G:G>100000))"
    
    ws['A15'] = "% Inactive Clients:"
    ws['B15'] = "=COUNTIF('Main Data'!E:E,0)/COUNTA('Main Data'!E:E)"
    ws['B15'].number_format = "0.00%"
    
    ws['A16'] = "Churn Risk Score:"
    ws['B16'] = "=IF(B15>0.2,\"HIGH\",IF(B15>0.1,\"MEDIUM\",\"LOW\"))"
    
    # Growth Projections
    ws['E12'] = "Growth Projections"
    ws['E12'].font = Font(bold=True)
    
    ws['E13'] = "Projected Monthly Growth (3%):"
    ws['F13'] = "=B6*1.03"
    ws['F13'].number_format = '"$"#,##0.00'
    
    ws['E14'] = "Projected Annual Growth (40%):"
    ws['F14'] = "=B6*1.4"
    ws['F14'].number_format = '"$"#,##0.00'
    
    ws['E15'] = "Break-even Point:"
    ws['F15'] = "=F13/30"
    ws['F15'].number_format = '"$"#,##0.00'
    
    # Data Quality Metrics
    ws['A18'] = "Data Quality Metrics"
    ws['A18'].font = Font(bold=True)
    
    ws['A19'] = "Records with Issues:"
    ws['B19'] = "=COUNTA('Quality Issues'!A:A)-1"
    
    ws['A20'] = "Data Quality Score:"
    ws['B20'] = "=1-(B19/COUNTA('Main Data'!A:A))"
    ws['B20'].number_format = "0.00%"
    
    ws['A21'] = "Completeness Ratio:"
    ws['B21'] = "=COUNTA('Main Data'!B:B)/COUNTA('Main Data'!A:A)"
    ws['B21'].number_format = "0.00%"
    
    # Financial Ratios
    ws['E18'] = "Financial Ratios"
    ws['E18'].font = Font(bold=True)
    
    ws['E19'] = "Active/Total Ratio:"
    ws['F19'] = "=SUM('Main Data'!E:E)/COUNTA('Main Data'!E:E)"
    ws['F19'].number_format = "0.00%"
    
    ws['E20'] = "Revenue Concentration:"
    ws['F20'] = "=LARGE('Main Data'!G:G,1)/SUM('Main Data'!G:G)"
    ws['F20'].number_format = "0.00%"
    
    ws['E21'] = "Top 10% Revenue Share:"
    ws['F21'] = "=SUMPRODUCT(LARGE('Main Data'!G:G,ROW(1:50)))/SUM('Main Data'!G:G)"
    ws['F21'].number_format = "0.00%"

def add_financial_modeling_sheet(wb):
    """Create financial modeling with advanced functions"""
    ws = wb.create_sheet("Financial Models", 8)
    
    # Title
    ws['A1'] = "Financial Modeling & Projections"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Scenario Analysis
    ws['A3'] = "Scenario Analysis"
    ws['A3'].font = Font(size=14, bold=True)
    
    # Input parameters
    ws['A5'] = "Input Parameters"
    ws['A5'].font = Font(bold=True)
    
    ws['A6'] = "Base Monthly Growth Rate:"
    ws['B6'] = 0.03
    ws['B6'].number_format = "0.00%"
    
    ws['A7'] = "Conservative Growth Rate:"
    ws['B7'] = 0.015
    ws['B7'].number_format = "0.00%"
    
    ws['A8'] = "Aggressive Growth Rate:"
    ws['B8'] = 0.05
    ws['B8'].number_format = "0.00%"
    
    ws['A9'] = "Churn Rate:"
    ws['B9'] = 0.02
    ws['B9'].number_format = "0.00%"
    
    # Scenario projections
    ws['D5'] = "12-Month Projections"
    ws['D5'].font = Font(bold=True)
    
    # Headers
    ws['D6'] = "Month"
    ws['E6'] = "Conservative"
    ws['F6'] = "Base Case"
    ws['G6'] = "Aggressive"
    
    for cell in ws['D6:G6'][0]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    # Monthly projections
    base_value = "='Advanced Analytics'!B6"
    for month in range(1, 13):
        row = 6 + month
        ws[f'D{row}'] = month
        ws[f'E{row}'] = f"={base_value}*POWER(1+B7-B9,{month})"  # Conservative
        ws[f'F{row}'] = f"={base_value}*POWER(1+B6-B9,{month})"  # Base
        ws[f'G{row}'] = f"={base_value}*POWER(1+B8-B9,{month})"  # Aggressive
        
        for col in ['E', 'F', 'G']:
            ws[f'{col}{row}'].number_format = '"$"#,##0'
    
    # NPV Calculations
    ws['A11'] = "Net Present Value Analysis"
    ws['A11'].font = Font(bold=True)
    
    ws['A12'] = "Discount Rate:"
    ws['B12'] = 0.08
    ws['B12'].number_format = "0.00%"
    
    ws['A13'] = "NPV Conservative:"
    ws['B13'] = "=NPV(B12,E7:E18)"
    ws['B13'].number_format = '"$"#,##0'
    
    ws['A14'] = "NPV Base Case:"
    ws['B14'] = "=NPV(B12,F7:F18)"
    ws['B14'].number_format = '"$"#,##0'
    
    ws['A15'] = "NPV Aggressive:"
    ws['B15'] = "=NPV(B12,G7:G18)"
    ws['B15'].number_format = '"$"#,##0'
    
    # Monte Carlo Simulation Setup
    ws['A17'] = "Monte Carlo Simulation Inputs"
    ws['A17'].font = Font(bold=True)
    
    ws['A18'] = "Number of Simulations:"
    ws['B18'] = 1000
    
    ws['A19'] = "Growth Rate Mean:"
    ws['B19'] = 0.03
    ws['B19'].number_format = "0.00%"
    
    ws['A20'] = "Growth Rate Std Dev:"
    ws['B20'] = 0.01
    ws['B20'].number_format = "0.00%"
    
    # Statistical Analysis
    ws['D11'] = "Statistical Analysis"
    ws['D11'].font = Font(bold=True)
    
    ws['D12'] = "Best Case (95th percentile):"
    ws['E12'] = "=PERCENTILE(G7:G18,0.95)"
    ws['E12'].number_format = '"$"#,##0'
    
    ws['D13'] = "Worst Case (5th percentile):"
    ws['E13'] = "=PERCENTILE(E7:E18,0.05)"
    ws['E13'].number_format = '"$"#,##0'
    
    ws['D14'] = "Probability of Growth:"
    ws['E14'] = "=COUNTIF(F7:F18,\">\"&F7)/COUNT(F7:F18)"
    ws['E14'].number_format = "0.00%"

def add_dynamic_charts_sheet(wb):
    """Create dynamic charts with advanced features"""
    ws = wb.create_sheet("Dynamic Charts", 9)
    
    # Title
    ws['A1'] = "Dynamic Charts & Visualizations"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:H1')
    
    # Portfolio distribution data
    ws['A3'] = "Portfolio Value Distribution"
    ws['A3'].font = Font(bold=True)
    
    ws['A4'] = "Range"
    ws['B4'] = "Count"
    ws['C4'] = "Percentage"
    
    ranges = [
        ("$0 - $500", "=COUNTIFS('Main Data'!G:G,\">=0\",'Main Data'!G:G,\"<50000\")"),
        ("$500 - $1,000", "=COUNTIFS('Main Data'!G:G,\">=50000\",'Main Data'!G:G,\"<100000\")"),
        ("$1,000 - $2,000", "=COUNTIFS('Main Data'!G:G,\">=100000\",'Main Data'!G:G,\"<200000\")"),
        ("$2,000 - $5,000", "=COUNTIFS('Main Data'!G:G,\">=200000\",'Main Data'!G:G,\"<500000\")"),
        ("$5,000+", "=COUNTIF('Main Data'!G:G,\">=500000\")")
    ]
    
    for i, (range_label, formula) in enumerate(ranges, 5):
        ws[f'A{i}'] = range_label
        ws[f'B{i}'] = formula
        ws[f'C{i}'] = f"=B{i}/SUM(B5:B9)"
        ws[f'C{i}'].number_format = "0.00%"
    
    # Create bar chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Client Portfolio Distribution"
    chart.y_axis.title = 'Number of Clients'
    chart.x_axis.title = 'Portfolio Value Range'
    
    data = Reference(ws, min_col=2, min_row=4, max_row=9, max_col=2)
    cats = Reference(ws, min_col=1, min_row=5, max_row=9)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    chart.height = 10
    chart.width = 15
    ws.add_chart(chart, "E3")
    
    # Time series analysis
    ws['A12'] = "Monthly Trend Analysis"
    ws['A12'].font = Font(bold=True)
    
    # Generate monthly data
    ws['A13'] = "Month"
    ws['B13'] = "New Clients"
    ws['C13'] = "Revenue"
    ws['D13'] = "Cumulative"
    
    for month in range(1, 13):
        row = 13 + month
        ws[f'A{row}'] = f"Month {month}"
        ws[f'B{row}'] = f"=RANDBETWEEN(10,50)"
        ws[f'C{row}'] = f"=B{row}*RANDBETWEEN(500,2000)*100"
        ws[f'D{row}'] = f"=SUM(C14:C{row})" if row > 14 else f"=C{row}"
    
    # Line chart for trends
    line_chart = LineChart()
    line_chart.title = "Monthly Revenue Trend"
    line_chart.style = 13
    line_chart.y_axis.title = 'Revenue ($)'
    line_chart.x_axis.title = 'Month'
    
    data = Reference(ws, min_col=3, min_row=13, max_row=25, max_col=4)
    cats = Reference(ws, min_col=1, min_row=14, max_row=25)
    line_chart.add_data(data, titles_from_data=True)
    line_chart.set_categories(cats)
    
    line_chart.height = 10
    line_chart.width = 15
    ws.add_chart(line_chart, "E15")

def add_conditional_formatting(wb):
    """Add advanced conditional formatting"""
    # Format Main Data sheet
    if "Main Data" in wb.sheetnames:
        ws = wb["Main Data"]
        
        # Balance column conditional formatting
        if ws.max_column >= 7:  # Assuming balance is in column G
            balance_range = f"G2:G{ws.max_row}"
            
            # Color scale for balances
            color_scale = ColorScaleRule(
                start_type='min', start_color='FF6B6B',
                mid_type='percentile', mid_value=50, mid_color='FFEB3B',
                end_type='max', end_color='4CAF50'
            )
            ws.conditional_formatting.add(balance_range, color_scale)
            
            # Data bars for visual representation
            data_bar = DataBarRule(
                start_type='min', start_value=0,
                end_type='max', end_value=None,
                color='4472C4', showValue=True
            )
            
        # Status column formatting
        if ws.max_column >= 5:  # Assuming status is in column E
            status_range = f"E2:E{ws.max_row}"
            
            # Highlight inactive clients
            inactive_rule = CellIsRule(
                operator='equal', formula=['0'],
                fill=PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
            )
            ws.conditional_formatting.add(status_range, inactive_rule)
            
            # Highlight active clients
            active_rule = CellIsRule(
                operator='equal', formula=['1'],
                fill=PatternFill(start_color='E6FFE6', end_color='E6FFE6', fill_type='solid')
            )
            ws.conditional_formatting.add(status_range, active_rule)

def add_data_validation(wb):
    """Add data validation and dropdown lists"""
    if "Advanced Analytics" in wb.sheetnames:
        ws = wb["Advanced Analytics"]
        
        # Add dropdown for risk assessment
        dv = DataValidation(type="list", formula1='"LOW,MEDIUM,HIGH"', showDropDown=True)
        dv.prompt = "Select risk level"
        dv.promptTitle = "Risk Assessment"
        ws.add_data_validation(dv)
        dv.add('B16')
        
    if "Financial Models" in wb.sheetnames:
        ws = wb["Financial Models"]
        
        # Percentage validation
        dv_percent = DataValidation(type="decimal", operator="between", formula1=0, formula2=1)
        dv_percent.prompt = "Enter a value between 0 and 1"
        dv_percent.promptTitle = "Percentage Input"
        ws.add_data_validation(dv_percent)
        for cell in ['B6', 'B7', 'B8', 'B9', 'B12', 'B19', 'B20']:
            dv_percent.add(cell)

def main():
    """Main function to enhance the Excel workbook"""
    print("Adding advanced Excel formulas and functions...")
    
    # Load existing workbook
    try:
        wb = openpyxl.load_workbook('outputs/data_analysis_workbook.xlsx')
    except FileNotFoundError:
        print("Error: data_analysis_workbook.xlsx not found. Please run create_excel_analysis.py first.")
        return
    
    # Add new advanced sheets
    add_advanced_analytics_sheet(wb)
    add_financial_modeling_sheet(wb)
    add_dynamic_charts_sheet(wb)
    
    # Add conditional formatting and validation
    add_conditional_formatting(wb)
    add_data_validation(wb)
    
    # Save enhanced workbook
    wb.save('outputs/data_analysis_workbook.xlsx')
    print("Advanced formulas and functions added successfully!")
    
    print(f"\nWorkbook now contains {len(wb.sheetnames)} sheets:")
    for sheet in wb.sheetnames:
        print(f"  - {sheet}")
    
    print("\nNew features added:")
    print("  * Advanced Analytics with KPIs and financial ratios")
    print("  * Financial modeling with scenario analysis")
    print("  * NPV calculations and statistical functions")
    print("  * Dynamic charts and visualizations")
    print("  * Conditional formatting for data visualization")
    print("  * Data validation and dropdown lists")
    print("  * Complex Excel formulas (SUMPRODUCT, COUNTIFS, NPV, etc.)")

if __name__ == "__main__":
    main()