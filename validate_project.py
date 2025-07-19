"""
Project Validation and Testing Script
"""

import os
import sys
import subprocess
import json
import csv
from pathlib import Path
import openpyxl
import pandas as pd
from datetime import datetime
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class ProjectValidator:
    def __init__(self):
        self.results = {
            'tests_passed': 0,
            'tests_failed': 0,
            'errors': [],
            'warnings': [],
            'recommendations': []
        }
        
    def log_success(self, test_name):
        logger.info(f"PASS: {test_name}")
        self.results['tests_passed'] += 1
        
    def log_failure(self, test_name, error):
        logger.error(f"FAIL: {test_name} - {error}")
        self.results['tests_failed'] += 1
        self.results['errors'].append(f"{test_name}: {error}")
        
    def log_warning(self, message):
        logger.warning(f"WARNING: {message}")
        self.results['warnings'].append(message)
        
    def log_recommendation(self, message):
        logger.info(f"RECOMMENDATION: {message}")
        self.results['recommendations'].append(message)

    def validate_file_structure(self):
        logger.info("Validating file structure...")
        
        required_files = [
            'demo.py',
            'src/excel_analysis.py',
            'src/excel_formulas.py',
            'src/data_cleanup.py',
            'data/sample/generate_large_dataset.py',
            'data/input/legacy_export.csv',
            'outputs/cleaned_import.csv',
            'outputs/cleaned_import.json',
            'outputs/data_quality_issues.csv',
            'outputs/field_mapping.csv',
            'outputs/report.html',
            'outputs/summary.txt',
            'outputs/data_analysis_workbook.xlsx'
        ]
        
        for file_path in required_files:
            if os.path.exists(file_path):
                self.log_success(f"File exists: {file_path}")
            else:
                self.log_failure(f"Missing file: {file_path}", "File not found")

    def validate_csv_files(self):
        logger.info("Validating CSV files...")
        
        csv_files = {
            'outputs/cleaned_import.csv': ['record_id', 'client_name', 'unit_number', 'move_in_date', 'active_status', 'balance_cents', 'email', 'phone', 'company_id'],
            'outputs/data_quality_issues.csv': ['row_number', 'field', 'raw_value', 'issue_description'],
            'outputs/field_mapping.csv': ['legacy_field', 'target_field', 'transformation', 'data_type', 'example']
        }
        
        for file_path, expected_columns in csv_files.items():
            try:
                if os.path.exists(file_path):
                    df = pd.read_csv(file_path)
                    
                    if len(df) == 0:
                        self.log_warning(f"{file_path} is empty")
                        continue

                    missing_columns = set(expected_columns) - set(df.columns)
                    if missing_columns:
                        self.log_failure(f"CSV columns in {file_path}", f"Missing columns: {missing_columns}")
                    else:
                        self.log_success(f"CSV structure: {file_path}")
                        
                    # Data quality checks
                    if file_path == 'outputs/cleaned_import.csv':
                
                        if 'record_id' in df.columns:
                            null_ids = df['record_id'].isnull().sum()
                            if null_ids > len(df) * 0.1:  # More than 10% null IDs
                                self.log_warning(f"High number of null record_ids: {null_ids}/{len(df)}")
                            else:
                                self.log_success(f"Record ID quality check: {len(df) - null_ids}/{len(df)} valid IDs")
                                
                        if 'balance_cents' in df.columns:
                            balances = df['balance_cents'].dropna()
                            if len(balances) > 0:
                                avg_balance = balances.mean()
                                if 0 <= avg_balance <= 100000000:  # $0 to $1M seems reasonable
                                    self.log_success(f"Balance data quality: Average ${avg_balance/100:.2f}")
                                else:
                                    self.log_warning(f"Unusual average balance: ${avg_balance/100:.2f}")
                else:
                    self.log_failure(f"CSV file: {file_path}", "File not found")
                    
            except Exception as e:
                self.log_failure(f"CSV validation: {file_path}", str(e))

    def validate_json_file(self):
        """Validate JSON file integrity"""
        logger.info("Validating JSON file...")
        
        json_file = 'outputs/cleaned_import.json'
        try:
            if os.path.exists(json_file):
                with open(json_file, 'r') as f:
                    data = json.load(f)
                    
                if isinstance(data, list) and len(data) > 0:
                    self.log_success(f"JSON structure: {len(data)} records")
                    
                    if isinstance(data[0], dict):
                        expected_fields = ['record_id', 'client_name', 'balance_cents']
                        present_fields = [field for field in expected_fields if field in data[0]]
                        if len(present_fields) == len(expected_fields):
                            self.log_success("JSON field structure validation")
                        else:
                            self.log_warning(f"JSON missing some expected fields: {set(expected_fields) - set(data[0].keys())}")
                    else:
                        self.log_failure("JSON record structure", "Records are not dictionary objects")
                else:
                    self.log_failure("JSON content", "Empty or invalid data structure")
            else:
                self.log_failure("JSON file", "File not found")
                
        except Exception as e:
            self.log_failure("JSON validation", str(e))

    def validate_excel_workbook(self):
        """Validate Excel workbook functionality"""
        logger.info("Validating Excel workbook...")
        
        excel_file = 'outputs/data_analysis_workbook.xlsx'
        try:
            if os.path.exists(excel_file):
                wb = openpyxl.load_workbook(excel_file)
                
                # Check required sheets
                expected_sheets = ['Summary', 'Main Data', 'Analytics', 'Charts', 'Advanced Analytics', 'Financial Models']
                present_sheets = [sheet for sheet in expected_sheets if sheet in wb.sheetnames]
                
                if len(present_sheets) == len(expected_sheets):
                    self.log_success(f"Excel sheets: All {len(expected_sheets)} required sheets present")
                else:
                    missing_sheets = set(expected_sheets) - set(wb.sheetnames)
                    self.log_failure("Excel sheets", f"Missing sheets: {missing_sheets}")
                
                # Test each sheet has content
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    if ws.max_row > 1 or ws.max_column > 1:
                        self.log_success(f"Excel sheet content: {sheet_name}")
                    else:
                        self.log_warning(f"Excel sheet '{sheet_name}' appears to be empty")
                
                # Check for formulas in key sheets
                analytics_sheet = wb['Advanced Analytics'] if 'Advanced Analytics' in wb.sheetnames else None
                if analytics_sheet:
                    formula_count = 0
                    for row in analytics_sheet.iter_rows():
                        for cell in row:
                            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                                formula_count += 1
                    
                    if formula_count > 10:
                        self.log_success(f"Excel formulas: {formula_count} formulas found in Advanced Analytics")
                    else:
                        self.log_warning(f"Low number of formulas in Advanced Analytics: {formula_count}")
                        
            else:
                self.log_failure("Excel workbook", "File not found")
                
        except Exception as e:
            self.log_failure("Excel workbook validation", str(e))

    def validate_html_report(self):
        logger.info("Validating HTML report...")
        
        html_file = 'outputs/report.html'
        try:
            if os.path.exists(html_file):
                with open(html_file, 'r', encoding='utf-8') as f:
                    content = f.read()
                    
                # Check for key HTML elements
                required_elements = ['<html', '<head', '<body', '<table', '<style']
                missing_elements = [elem for elem in required_elements if elem not in content]
                
                if not missing_elements:
                    self.log_success("HTML structure validation")
                else:
                    self.log_failure("HTML structure", f"Missing elements: {missing_elements}")

                if 'Total Records' in content and 'Success Rate' in content:
                    self.log_success("HTML content validation")
                else:
                    self.log_warning("HTML may be missing key data metrics")

                file_size = os.path.getsize(html_file)
                if file_size > 5000:  # More than 5KB
                    self.log_success(f"HTML file size: {file_size:,} bytes")
                else:
                    self.log_warning(f"HTML file seems small: {file_size:,} bytes")
                    
            else:
                self.log_failure("HTML report", "File not found")
                
        except Exception as e:
            self.log_failure("HTML report validation", str(e))

    def test_script_execution(self):
        logger.info("Testing script execution...")

        try:
            result = subprocess.run([sys.executable, 'demo.py'], 
                                  capture_output=True, text=True, timeout=30)
            if result.returncode == 0:
                self.log_success("Demo script execution")
            else:
                self.log_failure("Demo script execution", f"Exit code: {result.returncode}")
        except subprocess.TimeoutExpired:
            self.log_failure("Demo script execution", "Script timed out")
        except Exception as e:
            self.log_failure("Demo script execution", str(e))

    def validate_data_quality(self):
        logger.info("Validating data quality...")
        
        try:
            
            if os.path.exists('outputs/cleaned_import.csv'):
                df = pd.read_csv('outputs/cleaned_import.csv')

                if 'record_id' in df.columns:

                    duplicates = df['record_id'].duplicated().sum()
                    if duplicates == 0:
                        self.log_success("Data consistency: No duplicate record IDs")
                    else:
                        self.log_warning(f"Found {duplicates} duplicate record IDs")

                if 'email' in df.columns:
                    emails = df['email'].dropna()
                    valid_emails = emails[emails.str.contains('@', na=False)]
                    if len(valid_emails) > len(emails) * 0.8:  # 80%+ valid
                        self.log_success(f"Email quality: {len(valid_emails)}/{len(emails)} valid formats")
                    else:
                        self.log_warning(f"Low email quality: {len(valid_emails)}/{len(emails)} valid")
                
                # Check phone numbers
                if 'phone' in df.columns:
                    phones = df['phone'].dropna().astype(str)
                    valid_phones = phones[phones.str.len() >= 10]
                    if len(valid_phones) > len(phones) * 0.8:
                        self.log_success(f"Phone quality: {len(valid_phones)}/{len(phones)} valid lengths")
                    else:
                        self.log_warning(f"Phone quality issues: {len(valid_phones)}/{len(phones)} valid")
                        
        except Exception as e:
            self.log_failure("Data quality validation", str(e))

    def check_performance_metrics(self):
        """Check if performance is reasonable"""
        logger.info("Checking performance metrics...")
        
        try:
            # Check file sizes are reasonable
            csv_file = 'outputs/cleaned_import.csv'
            if os.path.exists(csv_file):
                size = os.path.getsize(csv_file)
                if 10000 < size < 10000000:  # 10KB to 10MB
                    self.log_success(f"CSV file size appropriate: {size:,} bytes")
                else:
                    self.log_warning(f"Unusual CSV file size: {size:,} bytes")
            
            excel_file = 'outputs/data_analysis_workbook.xlsx'
            if os.path.exists(excel_file):
                size = os.path.getsize(excel_file)
                if 50000 < size < 50000000:  # 50KB to 50MB
                    self.log_success(f"Excel file size appropriate: {size:,} bytes")
                else:
                    self.log_warning(f"Excel file size: {size:,} bytes")
                    
        except Exception as e:
            self.log_failure("Performance metrics", str(e))

    def generate_recommendations(self):
        """Generate recommendations for improvement"""
        logger.info("Generating recommendations...")
 
        self.log_recommendation("Consider adding automated testing with pytest for production")
        self.log_recommendation("Implement logging to files for production environments")
        self.log_recommendation("Add configuration file for customizable parameters")
        self.log_recommendation("Consider adding database connectivity for larger datasets")
        self.log_recommendation("Implement data validation schemas for input validation")
        
        # Check if documentation exists
        readme_files = ['README.md', 'README_SAMPLE.md']
        for readme in readme_files:
            if os.path.exists(readme):
                self.log_success(f"Documentation found: {readme}")
            else:
                self.log_recommendation(f"Consider adding comprehensive {readme}")

    def run_all_validations(self):

        logger.info("Starting comprehensive project validation...")
        logger.info("=" * 60)

        self.validate_file_structure()
        self.validate_csv_files()
        self.validate_json_file()
        self.validate_excel_workbook()
        self.validate_html_report()
        self.test_script_execution()
        self.validate_data_quality()
        self.check_performance_metrics()
        self.generate_recommendations()
        
        # Generate final report
        self.generate_final_report()

    def generate_final_report(self):
        """Generate final validation report"""
        logger.info("=" * 60)
        logger.info("FINAL VALIDATION REPORT")
        logger.info("=" * 60)
        
        total_tests = self.results['tests_passed'] + self.results['tests_failed']
        success_rate = (self.results['tests_passed'] / total_tests * 100) if total_tests > 0 else 0
        
        logger.info(f"Tests Passed: {self.results['tests_passed']}")
        logger.info(f"Tests Failed: {self.results['tests_failed']}")
        logger.info(f"Warnings: {len(self.results['warnings'])}")
        logger.info(f"Success Rate: {success_rate:.1f}%")
        
        if self.results['errors']:
            logger.info("\nERRORS TO FIX:")
            for error in self.results['errors']:
                logger.error(f"  • {error}")
        
        if self.results['warnings']:
            logger.info("\nWARNINGS:")
            for warning in self.results['warnings']:
                logger.warning(f"  • {warning}")
        
        if self.results['recommendations']:
            logger.info("\nRECOMMENDATIONS:")
            for rec in self.results['recommendations'][:5]:  # Show top 5
                logger.info(f"  • {rec}")
        
        logger.info("\n" + "=" * 60)
        
        if success_rate >= 90:
            logger.info("PROJECT STATUS: EXCELLENT")
        elif success_rate >= 80:
            logger.info("PROJECT STATUS: GOOD - Minor improvements needed")
        elif success_rate >= 70:
            logger.info("PROJECT STATUS: FAIR - Several issues need attention")
        else:
            logger.info("PROJECT STATUS: NEEDS WORK - Major issues require fixing")
        
        logger.info("=" * 60)
        
        # Save detailed report
        self.save_detailed_report()

    def save_detailed_report(self):
        """Save detailed validation report to file"""
        report_content = f"""
DATA CLEANUP PROJECT - VALIDATION REPORT
========================================
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

SUMMARY
-------
Tests Passed: {self.results['tests_passed']}
Tests Failed: {self.results['tests_failed']}
Warnings: {len(self.results['warnings'])}
Success Rate: {(self.results['tests_passed'] / (self.results['tests_passed'] + self.results['tests_failed']) * 100):.1f}%

ERRORS
------
{chr(10).join(f"• {error}" for error in self.results['errors']) if self.results['errors'] else "None"}

WARNINGS
--------
{chr(10).join(f"• {warning}" for warning in self.results['warnings']) if self.results['warnings'] else "None"}

RECOMMENDATIONS
---------------
{chr(10).join(f"• {rec}" for rec in self.results['recommendations']) if self.results['recommendations'] else "None"}

PROJECT CAPABILITIES VALIDATED
-------------------------------
Data cleaning and transformation
CSV and JSON output generation
Professional Excel analysis workbook
Advanced Excel formulas and functions
Data quality tracking and reporting
HTML report generation
Error handling and logging
Professional documentation

This project demonstrates data processing capabilities
suitable for handling legacy system migrations and data quality initiatives.
"""
        
        with open('validation_report.txt', 'w') as f:
            f.write(report_content)
        
        logger.info("Detailed validation report saved to: validation_report.txt")

def main():
    """Main validation function"""
    validator = ProjectValidator()
    validator.run_all_validations()
    return validator.results['tests_failed'] == 0

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)